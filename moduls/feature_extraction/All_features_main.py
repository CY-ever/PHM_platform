from moduls.feature_extraction.frequency_feature.All_frequency_features import *
import scipy.io as sio
import numpy as np
import copy
import os
from moduls.feature_extraction.time_domain_feature.writein import writein
import pandas as pd
from flask import abort

import openpyxl
import csv

from utils.table_setting import *
from utils.save_data import save_data_func

from moduls.feature_extraction.Report_feature_extraction import *


def features_save_data_func(data,names,output_file,save_path,file_name):
    '''

    :param data: 输入数据
    :param output_file: 保存格式
    :param save_path: 保存路径
    :param file_name: 文件名，同时也是mat文件的key名称
    :param index_label: excel和csv文件里的表头名
    :return:
    '''

# 选择保存文件的类型
    if output_file == 0:
        file_name=file_name+'.mat'
        path = os.path.join(save_path, file_name)
        sio.savemat(path, {file_name: data})
    elif output_file == 1:
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            # 设置表头
            worksheet.cell(1,1,"Index")
            for i in range(len(names)):
                worksheet.cell(1, i+2, names[i])
            for i in range(data.shape[0]):
                worksheet.cell(i + 2, 1, i)
            # names=list(names)
            # worksheet.append(names)
            for i in range(data.shape[0]):
                for j in range(data.shape[1]):
                    worksheet.cell(i+2,j+2,data[i,j])
            # 保存工作簿
            path = os.path.join(save_path, file_name+'.xlsx')
            workbook.save(path)

        except:
            abort(400, f"ERROR: This sheet is too big! Your worksheet size is: {data.shape}. The maximum excel worksheet size is (1048576, 16384).")

    elif output_file == 2:
        file_name = file_name+'.npy'
        path = os.path.join(save_path, file_name)
        np.save(path, np.array(data))
    elif output_file == 3:
        try:
            path = os.path.join(save_path, file_name+".csv")
            with open(path, 'w',newline='') as file:
                # 创建 writer 对象
                writer = csv.writer(file)
                index_data=[i for i in range(data.shape[0])]
                index_data=np.array(index_data)
                index_data.reshape((-1,1))
                # 写入表头

                names=np.insert(names,0,"Index")
                data=np.insert(data, 0, values=index_data, axis=1)
                writer.writerow(names)
                writer.writerows(data)

        except:
            abort(400,f"ERROR: This sheet is too big! Your worksheet size is: {data.shape}. The maximum excel worksheet size is (1048576, 16384).")
    elif output_file == 4:
        file_name = file_name+".txt"
        path = os.path.join(save_path, file_name)
        np.savetxt(path, data)

def names_save_data_func(data,output_file,save_path,file_name):
    '''

    :param data: 输入数据
    :param output_file: 保存格式
    :param save_path: 保存路径
    :param file_name: 文件名，同时也是mat文件的key名称
    :param index_label: excel和csv文件里的表头名
    :return:
    '''

# 选择保存文件的类型
    if output_file == 0:
        file_name=file_name+'.mat'
        path = os.path.join(save_path, file_name)
        sio.savemat(path, {file_name: data})
    elif output_file == 1:
        try:
            dataframe = pd.DataFrame({"Feature names":data})
            file_name = file_name+'.xlsx'
            path = os.path.join(save_path, file_name)
            writer = pd.ExcelWriter(path)
            dataframe.to_excel(writer,index=True,index_label="Index")
            writer.save()
        except:
            abort(400,f"ERROR: This sheet is too big! Your worksheet size is: {data.shape}. The maximum excel worksheet size is (1048576, 16384).")

    elif output_file == 2:
        file_name = file_name+'.npy'
        path = os.path.join(save_path, file_name)
        np.save(path, np.array(data))
    elif output_file == 3:
        try:
            file_name = file_name+".csv"
            dataframe = pd.DataFrame({"Feature names":data})
            path = os.path.join(save_path, file_name)
            dataframe.to_csv(path,index=True, sep=',',index_label="Index")
        except:
            abort(400,f"ERROR: This sheet is too big! Your worksheet size is: {data.shape}. The maximum excel worksheet size is (1048576, 16384).")
    elif output_file == 4:
        file_name = file_name+".txt"
        path = os.path.join(save_path, file_name)
        np.savetxt(path, data,fmt="%s")


def All_feature_extraction(dataset, labels,t_features=0,f_features=1,t_normal_features=(1,1,1,1,1,1,1,1,1,1,1,1,1),f_features_list=(1,1,1,1,1,1,1,1,1), DWT_EnergyEntropy_name='db1', DWT_EnergyEntropy_N=3, DWT_SingularEntropy_name='db1',
    DWT_SingularEntropy_N=3, WaveletPacket_EnergyEntropy_name='db1', WaveletPacket_EnergyEntropy_N=5,
    WaveletPacket_SingularEntropy_name='db1', WaveletPacket_SingularEntropy_N=5, OPFCF_fault_type_list= (1, 0, 0, 0),
    OPFCF_fr=25, OPFCF_order=6, OPFCF_fs=10000, OPFCF_switch=1, OPFCF_delta_f0=10, OPFCF_threshold=3,
    OPFCF_k=3, OPFCF_n_ball=8, OPFCF_d_ball=7.92, OPFCF_d_pitch=34.55, OPFCF_alpha=0, OPFCF_RUL_image=True,
                           data_SBN0=None,
                                          f_normal_features=(1,1,1,1,1,1,1,1,1,1,1,1,1),emd_fr=25,
    emd_n_ball=16, emd_d_ball=22.225, emd_d_pitch=np.dot(0.5, (102.7938 + 147.7264)) , emd_alpha=0, emd_fs=10000, emd_fault_type=0, emd_n=3, emd_ord=3, emd_limit=2000,
    FCF_ratio_nlevel=4, FCF_ratio_order=8, FCF_ratio_fs=10000, FCF_ratio_fr=25, FCF_ratio_n_ball=8,
    FCF_ratio_d_ball=7.92, FCF_ratio_d_pitch=34.55, FCF_ratio_alpha=0,FCF_ratio_image=False, save_path='./',output_file=0,output_image=0):
    '''
    :param data_set:input data，输入可以为一维或者二维数组，例如(10000,)或（24x10000)
    :param labels:一般为二维数组，例如（24x1)，只在FCF-ratio中用到，也支持一维数组，在代码中会被自动转换为二维数组
    :param OPFCF_num：此参数为数据长度，即采样点个数
    :param output_file: type to save file,0:mat,1:xlsx,2:npy,3:csv,4:txt
    :return: 输出为特征，每一组数据样本最多包含了13个时域特征和21个频域特征
    '''
    # global dataset_t_features_all, dataset_f_features_all
    global t_features_selected, f_features_selected, dataset_t_features_all, dataset_f_features_all, time_name, f_name
    if dataset.ndim==1:
        dataset=dataset.reshape(1,-1)#若输入为一维数组，则被自动转换为二维
    if labels is not None and labels.ndim==1:
        labels = labels.reshape(-1, 1)

    #判断是否有时域及频域特征被选中
    if t_features==1 and any(t_normal_features)!=False:
        t_features_selected = 1
        dataset_t_features_all,time_name = ALL_time_features(dataset, t_normal_features)
    else:
        t_features_selected = 0

    if f_features==1 and any(f_features_list)!=False:
        f_features_selected = 1
        dataset_f_features_all,f_name,labels = ALL_frequency_features(dataset, labels, f_features_list, DWT_EnergyEntropy_name,
                                                    DWT_EnergyEntropy_N, DWT_SingularEntropy_name,
                                                    DWT_SingularEntropy_N, WaveletPacket_EnergyEntropy_name,
                                                    WaveletPacket_EnergyEntropy_N,
                                                    WaveletPacket_SingularEntropy_name,
                                                    WaveletPacket_SingularEntropy_N, OPFCF_fault_type_list,
                                                    OPFCF_fr, OPFCF_order,  OPFCF_fs, OPFCF_switch,
                                                    OPFCF_delta_f0, OPFCF_threshold,
                                                    OPFCF_k, OPFCF_n_ball, OPFCF_d_ball, OPFCF_d_pitch, OPFCF_alpha,
                                                    OPFCF_RUL_image,data_SBN0,
                                                    f_normal_features, emd_fr,
                                                    emd_n_ball, emd_d_ball, emd_d_pitch, emd_alpha, emd_fs,
                                                    emd_fault_type, emd_n, emd_ord, emd_limit,
                                                    FCF_ratio_nlevel, FCF_ratio_order, FCF_ratio_fs, FCF_ratio_fr,
                                                    FCF_ratio_n_ball,
                                                    FCF_ratio_d_ball, FCF_ratio_d_pitch, FCF_ratio_alpha,
                                                    FCF_ratio_image, save_path, output_image)
    else:
        f_features_selected = 0

    if t_features_selected == 1 and f_features_selected==1:
        features_all = []
        dataset_t_features_all = dataset_t_features_all.tolist()
        dataset_f_features_all = dataset_f_features_all.tolist()
        time_name.extend(f_name)
        name_all=np.array(time_name)#将列表转为数组，为了方便集成时在特征选择模块的输入为数组
        for i in range(dataset.shape[0]):
            dataset_t_features_all[i].extend(dataset_f_features_all[i])
            features_all.append(dataset_t_features_all[i])
        features_all = np.array(features_all)
        # 保存总的特征，特征名，label
        features_save_data_func(data=features_all, names=name_all,  output_file=output_file, save_path=save_path,
                       file_name="All_features")

        names_save_data_func(data=name_all, output_file=output_file, save_path=save_path,
                       file_name="All_feature_names")
        if labels is not None:
            #保存labels
            features_save_data_func(data=labels, names=['Labels'], output_file=output_file, save_path=save_path,
                                    file_name="Labels_after_feature_extraction")

        # 生成报告
        word_feature_extraction(inputdata=dataset,inputlabels=labels,outputdata=features_all,output_labels=labels, t_features_selected=t_features_selected,f_features_selected=f_features_selected,f_features_list=f_features_list,all_name=name_all,DWT_energy_p1=DWT_EnergyEntropy_name,DWT_energy_p2=DWT_EnergyEntropy_N,DWT_singular_p1=DWT_SingularEntropy_name,DWT_singular_p2=DWT_SingularEntropy_N,WP_energy_p1=WaveletPacket_EnergyEntropy_name,WP_energy_p2=WaveletPacket_EnergyEntropy_N,WP_singular_p1=WaveletPacket_SingularEntropy_name,WP_singular_p2=WaveletPacket_SingularEntropy_N,OPFCF_fault_type_list=OPFCF_fault_type_list,OPFCF_fr=OPFCF_fr, OPFCF_order=OPFCF_order, OPFCF_fs=OPFCF_fs, OPFCF_switch=OPFCF_switch, OPFCF_delta_f0=OPFCF_delta_f0, OPFCF_threshold=OPFCF_threshold, OPFCF_k=OPFCF_k, OPFCF_n_ball=OPFCF_n_ball, OPFCF_d_ball=OPFCF_d_ball, OPFCF_d_pitch=OPFCF_d_pitch, OPFCF_alpha=OPFCF_alpha, EMD_fr=emd_fr, EMD_n_ball=emd_n_ball, EMD_d_ball=emd_d_ball, EMD_d_pitch=emd_d_pitch, EMD_alpha=emd_alpha,EMD_fs=emd_fs,EMD_fault_type=emd_fault_type,EMD_n=emd_n,EMD_ord=emd_ord,EMD_limit=emd_limit, FCF_nlevel=FCF_ratio_nlevel, FCF_order=FCF_ratio_order, FCF_fs=FCF_ratio_fs, FCF_fr=FCF_ratio_fr, FCF_n_ball=FCF_ratio_n_ball, FCF_d_ball=FCF_ratio_d_ball, FCF_d_pitch=FCF_ratio_d_pitch, FCF_alpha=FCF_ratio_alpha,FCF_output_image=FCF_ratio_image,output_file=output_file,save_path=save_path)

        return features_all, labels, name_all

    elif t_features_selected == 1 and f_features_selected==0:
        features_all= dataset_t_features_all
        name_all = np.array(time_name)

        features_save_data_func(data=features_all, names=name_all,output_file=output_file, save_path=save_path,
                       file_name="Time_domain_features")
        names_save_data_func(data=name_all, output_file=output_file, save_path=save_path,
                       file_name="Time_domain_feature_names")

        if labels is not None:
            # 保存labels
            features_save_data_func(data=labels, names=['Labels'], output_file=output_file, save_path=save_path,
                                    file_name="Labels_after_feature_extraction")

        # 生成报告
        word_feature_extraction(inputdata=dataset, inputlabels=labels,outputdata=features_all, output_labels=labels,
                                t_features_selected=t_features_selected, f_features_selected=f_features_selected,
                                f_features_list=f_features_list, all_name=name_all,
                                DWT_energy_p1=DWT_EnergyEntropy_name, DWT_energy_p2=DWT_EnergyEntropy_N,
                                DWT_singular_p1=DWT_SingularEntropy_name, DWT_singular_p2=DWT_SingularEntropy_N,
                                WP_energy_p1=WaveletPacket_EnergyEntropy_name,
                                WP_energy_p2=WaveletPacket_EnergyEntropy_N,
                                WP_singular_p1=WaveletPacket_SingularEntropy_name,
                                WP_singular_p2=WaveletPacket_SingularEntropy_N,
                                OPFCF_fault_type_list=OPFCF_fault_type_list, OPFCF_fr=OPFCF_fr, OPFCF_order=OPFCF_order,
                                OPFCF_fs=OPFCF_fs, OPFCF_switch=OPFCF_switch, OPFCF_delta_f0=OPFCF_delta_f0,
                                OPFCF_threshold=OPFCF_threshold, OPFCF_k=OPFCF_k, OPFCF_n_ball=OPFCF_n_ball,
                                OPFCF_d_ball=OPFCF_d_ball, OPFCF_d_pitch=OPFCF_d_pitch, OPFCF_alpha=OPFCF_alpha,
                                EMD_fr=emd_fr, EMD_n_ball=emd_n_ball, EMD_d_ball=emd_d_ball, EMD_d_pitch=emd_d_pitch,
                                EMD_alpha=emd_alpha, EMD_fs=emd_fs, EMD_fault_type=emd_fault_type, EMD_n=emd_n,
                                EMD_ord=emd_ord, EMD_limit=emd_limit, FCF_nlevel=FCF_ratio_nlevel,
                                FCF_order=FCF_ratio_order, FCF_fs=FCF_ratio_fs, FCF_fr=FCF_ratio_fr,
                                FCF_n_ball=FCF_ratio_n_ball, FCF_d_ball=FCF_ratio_d_ball, FCF_d_pitch=FCF_ratio_d_pitch,
                                FCF_alpha=FCF_ratio_alpha, FCF_output_image=FCF_ratio_image, output_file=output_file,
                                save_path=save_path)

        return features_all, labels, name_all
    elif t_features_selected == 0 and f_features_selected==1:
        features_all = dataset_f_features_all
        name_all=np.array(f_name)
        features_save_data_func(data=features_all, names=name_all, output_file=output_file, save_path=save_path,
                       file_name="Frequency_features")
        names_save_data_func(data=name_all, output_file=output_file, save_path=save_path,
                       file_name="Frequency_feature_names")

        if labels is not None:
            # 保存labels
            features_save_data_func(data=labels, names=['Labels'], output_file=output_file, save_path=save_path,
                                    file_name="Labels_after_feature_extraction")

        # 生成报告
        word_feature_extraction(inputdata=dataset, inputlabels=labels,outputdata=features_all, output_labels=labels,
                                t_features_selected=t_features_selected, f_features_selected=f_features_selected,
                                f_features_list=f_features_list, all_name=name_all,
                                DWT_energy_p1=DWT_EnergyEntropy_name, DWT_energy_p2=DWT_EnergyEntropy_N,
                                DWT_singular_p1=DWT_SingularEntropy_name, DWT_singular_p2=DWT_SingularEntropy_N,
                                WP_energy_p1=WaveletPacket_EnergyEntropy_name,
                                WP_energy_p2=WaveletPacket_EnergyEntropy_N,
                                WP_singular_p1=WaveletPacket_SingularEntropy_name,
                                WP_singular_p2=WaveletPacket_SingularEntropy_N,
                                OPFCF_fault_type_list=OPFCF_fault_type_list, OPFCF_fr=OPFCF_fr, OPFCF_order=OPFCF_order,
                                OPFCF_fs=OPFCF_fs, OPFCF_switch=OPFCF_switch, OPFCF_delta_f0=OPFCF_delta_f0,
                                OPFCF_threshold=OPFCF_threshold, OPFCF_k=OPFCF_k, OPFCF_n_ball=OPFCF_n_ball,
                                OPFCF_d_ball=OPFCF_d_ball, OPFCF_d_pitch=OPFCF_d_pitch, OPFCF_alpha=OPFCF_alpha,
                                EMD_fr=emd_fr, EMD_n_ball=emd_n_ball, EMD_d_ball=emd_d_ball, EMD_d_pitch=emd_d_pitch,
                                EMD_alpha=emd_alpha, EMD_fs=emd_fs, EMD_fault_type=emd_fault_type, EMD_n=emd_n,
                                EMD_ord=emd_ord, EMD_limit=emd_limit, FCF_nlevel=FCF_ratio_nlevel,
                                FCF_order=FCF_ratio_order, FCF_fs=FCF_ratio_fs, FCF_fr=FCF_ratio_fr,
                                FCF_n_ball=FCF_ratio_n_ball, FCF_d_ball=FCF_ratio_d_ball, FCF_d_pitch=FCF_ratio_d_pitch,
                                FCF_alpha=FCF_ratio_alpha, FCF_output_image=FCF_ratio_image, output_file=output_file,
                                save_path=save_path)
        return features_all, labels, name_all
    elif t_features_selected == 0 and f_features_selected==0:
        name_all=None
        return dataset,labels,name_all #dataset和labels输出都为二维数组


def ALL_time_features(dataset,t_normal_features):
    global time_name
    dataset_t_features_all = []
    for i in range(dataset.shape[0]):
        input_signal=dataset[i]

        time_features=[]
        time_name=[]
        signal_size = np.size(input_signal)
        if t_normal_features[0]==1:
            time_features_max = np.max(input_signal)
            time_features.append(time_features_max)
            time_name.append('t_max')
        if t_normal_features[1] == 1:
            time_features_min = np.min(input_signal)
            time_features.append(time_features_min)
            time_name.append('t_min')
        if t_normal_features[2] == 1:
            time_features_mean = np.mean(input_signal)
            time_features.append(time_features_mean)
            time_name.append('t_mean')
        if t_normal_features[3] == 1:
            time_features_root_mean_square = np.sqrt(np.mean(input_signal ** 2))
            time_features.append(time_features_root_mean_square)
            time_name.append('t_root_mean_square')
        if t_normal_features[4] == 1:
            time_features_mean = np.mean(input_signal)
            time_features_standard_deviation = np.sqrt(
            np.sum((input_signal - time_features_mean) ** 2) / (signal_size - 1)) #此处原本是10001
            time_features.append(time_features_standard_deviation)
            time_name.append('t_standard_deviation')
        if t_normal_features[5] == 1:
            time_features_mean = np.mean(input_signal)
            time_features_variance = np.sum((input_signal - time_features_mean) ** 2) / (
                    signal_size - 1)
            time_features.append(time_features_variance)
            time_name.append('t_variance')
        if t_normal_features[6] == 1:
            time_features_median = np.median(input_signal)
            time_features.append(time_features_median)
            time_name.append('t_median')
        if t_normal_features[7] == 1:
            time_features_mean = np.mean(input_signal)
            time_features_standard_deviation = np.sqrt(
                np.sum((input_signal - time_features_mean) ** 2) / (signal_size - 1))
            time_features_skewness = np.sum(((input_signal - time_features_mean) /
            time_features_standard_deviation) ** 3) / signal_size
            time_features.append(time_features_skewness )
            time_name.append('t_skewness')
        if t_normal_features[8] == 1:
            time_features_mean = np.mean(input_signal)
            time_features_standard_deviation = np.sqrt(
                np.sum((input_signal - time_features_mean) ** 2) / (signal_size - 1))
            time_features_kurtosis = np.sum(((input_signal - time_features_mean) /
            time_features_standard_deviation) ** 4) / signal_size
            time_features.append(time_features_kurtosis)
            time_name.append('t_kurtosis')
        if t_normal_features[9] == 1:
            time_features_max = np.max(input_signal)
            time_features_min = np.min(input_signal)
            time_features_peak_to_peak_value = time_features_max - time_features_min
            time_features.append(time_features_peak_to_peak_value)
            time_name.append('t_peak_to_peak_value')
        if t_normal_features[10] == 1:
            time_features_root_mean_square = np.sqrt(np.mean(input_signal ** 2))
            time_features_crest_factor = np.max(np.abs(input_signal)) / time_features_root_mean_square
            time_features.append(time_features_crest_factor)
            time_name.append('t_crest_factor')
        if t_normal_features[11] == 1:
            time_features_root_mean_square = np.sqrt(np.mean(input_signal ** 2))
            time_features_shape_factor = time_features_root_mean_square / np.mean(np.abs(input_signal))
            time_features.append(time_features_shape_factor)
            time_name.append('t_shape_factor')
        if t_normal_features[12] == 1:
            time_features_impulse_factor = np.max(np.abs(input_signal)) / np.mean(np.abs(input_signal))
            time_features.append(time_features_impulse_factor)
            time_name.append('t_impulse_factor')
        dataset_t_features_all.append(time_features)
    dataset_t_features_all = np.array(dataset_t_features_all)


    return  dataset_t_features_all,time_name

if __name__=="__main__":
    signal=writein('./Frequency_features/3_segmentation_data_100_3000.mat',1)
    labels=None
    # labels = writein('./Frequency_features/3_segmentation_label_100_3000.mat',1)

    features,labels,name_all=All_feature_extraction(signal, labels,t_features=0,f_features=1,save_path="./result")





