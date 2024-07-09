from moduls.feature_selection.SVD import SVD
from moduls.feature_selection.PCA import PCA_main
from moduls.feature_selection.FDA import FDA
from moduls.feature_selection.AE import AE
from moduls.feature_selection.Monotonicity import evaluate_monotonicity
from moduls.feature_selection.Correlation import evaluate_correlation
from moduls.feature_selection.Report_feature_selection import *
from moduls.feature_selection.writein import writein
from utils.table_setting import *
from utils.save_data import save_data_func
from flask import abort
import openpyxl
import scipy.io as sio
import csv


#在这里用于保存labels
def features_save_data_func(data,names,output_file,save_path,file_name):
    '''

    :param data: 输入数据
    :param output_file: 保存格式
    :param save_path: 保存路径
    :param file_name: 文件名，同时也是mat文件的key名称
    :param index_label: excel和csv文件里的表头名
    :return:
    '''

# 选择保存文件的类型
    if output_file == 0:
        file_name=file_name+'.mat'
        path = os.path.join(save_path, file_name)
        sio.savemat(path, {file_name: data})
    elif output_file == 1:
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            # 设置表头
            worksheet.cell(1,1,"Index")
            for i in range(len(names)):
                worksheet.cell(1, i+2, names[i])
            for i in range(data.shape[0]):
                worksheet.cell(i + 2, 1, i)
            for i in range(data.shape[0]):
                for j in range(data.shape[1]):
                    worksheet.cell(i+2,j+2,data[i,j])
            # 保存工作簿
            path = os.path.join(save_path, file_name+'.xlsx')
            workbook.save(path)

        except:
            abort(400, f"ERROR: This sheet is too big! Your worksheet size is: {data.shape}. The maximum excel worksheet size is (1048576, 16384).")

    elif output_file == 2:
        file_name = file_name+'.npy'
        path = os.path.join(save_path, file_name)
        np.save(path, np.array(data))
    elif output_file == 3:
        try:
            path = os.path.join(save_path, file_name+".csv")
            with open(path, 'w',newline='') as file:
                # 创建 writer 对象
                writer = csv.writer(file)
                index_data=[i for i in range(data.shape[0])]
                index_data=np.array(index_data)
                index_data.reshape((-1,1))
                # 写入表头

                names=np.insert(names,0,"Index")
                data=np.insert(data, 0, values=index_data, axis=1)
                writer.writerow(names)
                writer.writerows(data)

        except:
            abort(400,f"ERROR: This sheet is too big! Your worksheet size is: {data.shape}. The maximum excel worksheet size is (1048576, 16384).")
    elif output_file == 4:
        file_name = file_name+".txt"
        path = os.path.join(save_path, file_name)
        np.savetxt(path, data)

def Features_Selection_main(data,labels,name_list,Features_selection=3,svd_dimension=2, pca_method=1,pca_dimension_method=0,pca_dimension=20,pca_percent=80,fda_dim=2,AE_encoding_dim=2,Monotonicity_threshold=0,Correlation_threshold=0,save_path='./',output_file=0,output_image=0):
    #保存labels
    if labels is not None:
        if labels.ndim==1:
            labels=labels.reshape((-1,1))
        else:
            pass
        features_save_data_func(data=labels, names=['Labels'], output_file=output_file, save_path=save_path,
                                file_name="Labels_after_feature_selection")

    if Features_selection==2:
        new_data=SVD(data, svd_dimension, save_path, output_file)

    elif Features_selection==3:
        new_data=PCA_main(data, pca_method,pca_dimension_method,pca_dimension,pca_percent, save_path,
                 output_file)
    elif Features_selection == 4:

        new_data=FDA(data, labels, fda_dim, save_path, output_file)

    elif Features_selection == 5:
        new_data=AE(data, AE_encoding_dim, save_path, output_file)

    elif Features_selection == 1:
        new_data=evaluate_monotonicity(data,name_list,Monotonicity_threshold,save_path,output_file,output_image)
    elif Features_selection == 0:
        new_data= evaluate_correlation(data, name_list,Correlation_threshold,save_path,output_file,output_image)
    #生成报告
    word_feature_selection(inputdata=data, outputdata=new_data, input_labels=labels,output_labels=labels, name_list=name_list, Features_selection=Features_selection, svd_dimension=svd_dimension,
                           pca_method=pca_method, pca_dimension_method=pca_dimension_method, pca_dimension=pca_dimension, pca_percent=pca_percent, fda_dim=fda_dim, AE_encoding_dim=AE_encoding_dim,
                           Monotonicity_threshold=Monotonicity_threshold, Correlation_threshold=Correlation_threshold, output_file=output_file, output_image=output_image, save_path=save_path)
    return new_data,labels

if __name__=="__main__":
    labels=None
    data=writein('features_all.mat',1)
    # data=writein('3_segmentation_data_100_3000.mat',1)
    # labels = writein('3_segmentation_label_100_3000.mat',1)
    name_list=writein('name_all.mat',1)
    # name_list=None
    new_data = Features_Selection_main(data, labels,name_list,Features_selection=2,save_path="./",output_file=0)

